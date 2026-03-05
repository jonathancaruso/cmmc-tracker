#!/usr/bin/env python3
"""CMMC Artifact Tracker — Flask + SQLite"""

import os
import re
import sqlite3
import csv
import io
import time
import collections
from datetime import datetime
from flask import Flask, render_template, jsonify, request, Response, redirect, url_for, session, abort
from openpyxl import load_workbook
import functools
import secrets
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)

# --- Secret key: prefer env var; fallback generates random (invalidates sessions on restart) ---
_secret = os.environ.get("FLASK_SECRET")
if not _secret:
    _secret = secrets.token_hex(32)
    import sys
    print("WARNING: FLASK_SECRET not set. Using random key — sessions will not survive restarts. "
          "Set FLASK_SECRET env var for production.", file=sys.stderr)
app.secret_key = _secret

# --- Session / cookie security ---
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.environ.get("FLASK_ENV") == "production"
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100 MB upload limit

DB_PATH = os.environ.get("DB_PATH", os.path.join(os.path.dirname(__file__), "cmmc.db"))
XLSX_PATH = os.path.join(os.path.dirname(__file__), "nist-800-171a.xlsx")
XLSX_171_PATH = os.path.join(os.path.dirname(__file__), "nist-800-171.xlsx")
UPLOAD_DIR = os.environ.get("UPLOAD_PATH", os.path.join(os.path.dirname(__file__), "uploads"))
os.makedirs(UPLOAD_DIR, exist_ok=True)
ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.pdf', '.doc', '.docx',
                      '.xls', '.xlsx', '.txt', '.csv', '.pptx', '.zip', '.md'}


# --- Rate limiting for login ---
_login_attempts = collections.defaultdict(list)  # ip -> [timestamp, ...]
LOGIN_RATE_LIMIT = 5       # max attempts
LOGIN_RATE_WINDOW = 300    # per 5 minutes

def _check_rate_limit(ip):
    now = time.time()
    _login_attempts[ip] = [t for t in _login_attempts[ip] if now - t < LOGIN_RATE_WINDOW]
    if len(_login_attempts[ip]) >= LOGIN_RATE_LIMIT:
        return False
    _login_attempts[ip].append(now)
    return True


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    # Create all tables
    conn.execute("""
        CREATE TABLE IF NOT EXISTS objectives (
            id TEXT PRIMARY KEY,
            family TEXT NOT NULL,
            requirement_id TEXT NOT NULL,
            sort_as TEXT,
            security_requirement TEXT,
            assessment_objective TEXT,
            examine TEXT,
            interview TEXT,
            test TEXT,
            captured INTEGER DEFAULT 0,
            artifact_notes TEXT DEFAULT '',
            captured_date TEXT,
            requirement_type TEXT DEFAULT '',
            discussion TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS artifacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            file_size INTEGER,
            mime_type TEXT,
            uploaded_at TEXT NOT NULL,
            file_created TEXT,
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            role TEXT DEFAULT '',
            email TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS domains (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            color TEXT NOT NULL DEFAULT '#6366f1'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS artifact_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL,
            member_id INTEGER NOT NULL,
            status TEXT DEFAULT 'assigned',
            due_date TEXT,
            assigned_at TEXT NOT NULL,
            FOREIGN KEY (objective_id) REFERENCES objectives(id),
            FOREIGN KEY (member_id) REFERENCES team_members(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            first_name TEXT DEFAULT '',
            last_name TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS poam (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL UNIQUE,
            weakness TEXT DEFAULT '',
            remediation TEXT DEFAULT '',
            resources TEXT DEFAULT '',
            milestone_date TEXT,
            risk_level TEXT DEFAULT 'Moderate',
            updated_at TEXT NOT NULL,
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            action TEXT NOT NULL,
            target_type TEXT,
            target_id TEXT,
            details TEXT,
            timestamp TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS objective_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL,
            user_id INTEGER,
            username TEXT NOT NULL,
            comment TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS artifact_objectives (
            artifact_id INTEGER NOT NULL,
            objective_id TEXT NOT NULL,
            PRIMARY KEY (artifact_id, objective_id),
            FOREIGN KEY (artifact_id) REFERENCES artifacts(id),
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.commit()

    # Migrate existing DBs: add columns if they don't exist
    for col in ["requirement_type", "discussion"]:
        try:
            conn.execute(f"SELECT {col} FROM objectives LIMIT 1")
        except Exception:
            try:
                conn.execute(f"ALTER TABLE objectives ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass

    # Add status column
    try:
        conn.execute("SELECT status FROM objectives LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE objectives ADD COLUMN status TEXT DEFAULT 'Not Started'")
            # Backfill: set captured objectives to 'Complete'
            conn.execute("UPDATE objectives SET status = 'Complete' WHERE captured = 1")
        except Exception:
            pass
    # Migrate: add domain_id to artifacts
    try:
        conn.execute("SELECT domain_id FROM artifacts LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE artifacts ADD COLUMN domain_id INTEGER REFERENCES domains(id)")
        except Exception:
            pass

    # Migrate: add file_created to artifacts
    try:
        conn.execute("SELECT file_created FROM artifacts LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE artifacts ADD COLUMN file_created TEXT")
        except Exception:
            pass

    # Migrate: add first_name, last_name to users
    for col in ["first_name", "last_name"]:
        try:
            conn.execute(f"SELECT {col} FROM users LIMIT 1")
        except Exception:
            try:
                conn.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass

    # Migrate: add obtained_method to artifacts
    try:
        conn.execute("SELECT obtained_method FROM artifacts LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE artifacts ADD COLUMN obtained_method TEXT DEFAULT ''")
        except Exception:
            pass

    # No default domains — user adds their AD domains via config page

    # Migrate: populate artifact_objectives junction table from artifacts.objective_id
    try:
        existing = conn.execute("SELECT COUNT(*) FROM artifact_objectives").fetchone()[0]
        if existing == 0:
            conn.execute("""
                INSERT OR IGNORE INTO artifact_objectives (artifact_id, objective_id)
                SELECT id, objective_id FROM artifacts WHERE objective_id IS NOT NULL
            """)
    except Exception:
        pass

    conn.commit()

    # Check if objectives already seeded
    count = conn.execute("SELECT COUNT(*) FROM objectives").fetchone()[0]
    if count > 0:
        # Seed discussion if missing
        if os.path.exists(XLSX_171_PATH):
            sample = conn.execute("SELECT discussion FROM objectives WHERE requirement_id = '3.1.1' LIMIT 1").fetchone()
            if sample and not sample["discussion"]:
                wb2 = load_workbook(XLSX_171_PATH, read_only=True)
                ws2 = wb2["SP 800-171"]
                for row in ws2.iter_rows(min_row=2, values_only=True):
                    req_id = str(row[2] or "").strip()
                    req_type = str(row[1] or "").strip()
                    discussion = str(row[5] or "").strip()
                    if req_id:
                        conn.execute(
                            "UPDATE objectives SET requirement_type = ?, discussion = ? WHERE requirement_id = ?",
                            (req_type, discussion, req_id)
                        )
                conn.commit()
                wb2.close()
        conn.close()
        return

    # Parse xlsx — two passes: collect all rows, then insert
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb["SP800-171A"]
    current_family = ""
    current_req_id = ""
    current_sec_req = ""
    current_examine = ""
    current_interview = ""
    current_test = ""

    # Track which requirements have bracket sub-items
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    reqs_with_brackets = set()
    for row in all_rows:
        ident = str(row[1] or "").strip()
        if "[" in ident:
            reqs_with_brackets.add(ident.split("[")[0])

    for row in all_rows:
        family = row[0] or current_family
        identifier = str(row[1] or "").strip()
        sort_as = str(row[2] or "").strip()
        sec_req = row[3]
        obj_text = row[4]
        examine = row[5]
        interview = row[6]
        test = row[7]

        if not identifier:
            continue

        current_family = family

        if "[" not in identifier:
            # Parent requirement row — save context for sub-items
            current_req_id = identifier
            current_sec_req = sec_req or ""
            current_examine = examine or ""
            current_interview = interview or ""
            current_test = test or ""
            current_sort_as = sort_as
            current_obj_text = obj_text or ""
            # If this requirement has no bracket sub-items, insert it as a single objective
            if identifier not in reqs_with_brackets and current_obj_text:
                # Strip "Determine if:" or "Determine If: " prefix
                obj_clean = current_obj_text
                for prefix in ("Determine if: ", "Determine If: ", "Determine if:", "Determine If:"):
                    if obj_clean.startswith(prefix):
                        obj_clean = obj_clean[len(prefix):].strip()
                        break
                conn.execute("""
                    INSERT OR IGNORE INTO objectives
                    (id, family, requirement_id, sort_as, security_requirement,
                     assessment_objective, examine, interview, test)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    identifier, current_family, identifier, sort_as,
                    current_sec_req, obj_clean,
                    current_examine, current_interview, current_test
                ))
        else:
            # Assessment objective row (bracketed sub-item)
            conn.execute("""
                INSERT OR IGNORE INTO objectives 
                (id, family, requirement_id, sort_as, security_requirement, 
                 assessment_objective, examine, interview, test)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                identifier, current_family, current_req_id, sort_as,
                current_sec_req, obj_text or "",
                current_examine, current_interview, current_test
            ))

    conn.commit()
    wb.close()
    print(f"Database seeded from {XLSX_PATH}")

    # Now seed discussion and requirement_type from SP 800-171
    if os.path.exists(XLSX_171_PATH):
        wb2 = load_workbook(XLSX_171_PATH, read_only=True)
        ws2 = wb2["SP 800-171"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            req_id = str(row[2] or "").strip()
            req_type = str(row[1] or "").strip()
            sec_req = str(row[4] or "").strip()
            discussion = str(row[5] or "").strip()
            if req_id:
                conn.execute(
                    "UPDATE objectives SET requirement_type = ?, discussion = ? WHERE requirement_id = ?",
                    (req_type, discussion, req_id)
                )
                # Backfill missing security_requirement text
                if sec_req:
                    import re
                    sec_req_clean = re.sub(r'\[\d+\]\.?', '', sec_req).strip()
                    conn.execute(
                        "UPDATE objectives SET security_requirement = ? WHERE requirement_id = ? AND (security_requirement IS NULL OR security_requirement = '')",
                        (sec_req_clean, req_id)
                    )
        conn.commit()
        wb2.close()
        print(f"Discussion and requirement types seeded from {XLSX_171_PATH}")

    conn.close()


def validate_password(password):
    errors = []
    if len(password) < 16:
        errors.append("Password must be at least 16 characters")
    if not re.search(r'[A-Z]', password):
        errors.append("Must contain an uppercase letter")
    if not re.search(r'[a-z]', password):
        errors.append("Must contain a lowercase letter")
    if not re.search(r'[0-9]', password):
        errors.append("Must contain a number")
    if not re.search(r'[^A-Za-z0-9]', password):
        errors.append("Must contain a special character")
    return errors


def log_audit(action, target_type=None, target_id=None, details=None, conn=None):
    close = False
    if conn is None:
        conn = get_db()
        close = True
    conn.execute(
        "INSERT INTO audit_log (user_id, username, action, target_type, target_id, details, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (session.get('user_id'), session.get('username'), action, target_type,
         str(target_id) if target_id is not None else None, details,
         datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    if close:
        conn.close()


def admin_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        user_id = session.get('user_id')
        if not user_id:
            return redirect(url_for('login'))
        conn = get_db()
        user = conn.execute("SELECT role FROM users WHERE id = ?", (user_id,)).fetchone()
        conn.close()
        if not user or user['role'] != 'admin':
            if request.path.startswith('/api/'):
                return jsonify({"error": "Admin access required"}), 403
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


@app.before_request
def check_auth():
    if request.endpoint in ('static', None):
        return
    PUBLIC_ENDPOINTS = {'login', 'setup', 'landing'}
    conn = get_db()
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()
    if user_count == 0:
        if request.endpoint != 'setup':
            return redirect(url_for('setup'))
        return
    if request.endpoint in PUBLIC_ENDPOINTS:
        if 'user_id' in session:
            return redirect(url_for('dashboard'))
        return
    if 'user_id' not in session:
        if request.path.startswith('/api/') or request.path.startswith('/uploads/'):
            return jsonify({"error": "Authentication required"}), 401
        return redirect(url_for('login'))


# --- CSRF protection ---
@app.before_request
def csrf_protect():
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return
    if request.endpoint in ('static', None):
        return
    # Exempt logout (GET-based, session is cleared)
    if request.endpoint == 'logout':
        return
    # Form-based endpoints: validate CSRF from form field
    if request.endpoint in ('login', 'setup'):
        token = request.form.get('csrf_token', '')
        if not token or token != session.get('csrf_token'):
            if request.endpoint == 'login':
                return render_template("login.html", error="Session expired. Please try again.")
            abort(403)
        return
    # For API endpoints, validate CSRF header (also check form field for multipart uploads)
    token = request.headers.get('X-CSRF-Token', '') or request.form.get('csrf_token', '')
    if not token or token != session.get('csrf_token'):
        return jsonify({"error": "CSRF token missing or invalid"}), 403


def generate_csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(32)
    return session['csrf_token']


@app.context_processor
def inject_csrf():
    return {'csrf_token': generate_csrf_token}


# --- Security headers ---
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline'; "
        "style-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: blob:; "
        "font-src 'self'; "
        "form-action 'self'; "
        "frame-ancestors 'none';"
    )
    if os.environ.get("FLASK_ENV") == "production":
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response


# --- Error handlers ---
@app.errorhandler(404)
def not_found(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Not found"}), 404
    return "Page not found", 404


@app.errorhandler(413)
def too_large(e):
    return jsonify({"error": "File too large. Maximum size is 100 MB."}), 413


@app.errorhandler(500)
def internal_error(e):
    if request.path.startswith('/api/'):
        return jsonify({"error": "Internal server error"}), 500
    return "Internal server error", 500


@app.context_processor
def inject_user():
    if 'user_id' in session:
        conn = get_db()
        user = conn.execute("SELECT id, username, role, first_name, last_name FROM users WHERE id = ?",
                            (session['user_id'],)).fetchone()
        conn.close()
        if user:
            return {'current_user': dict(user)}
    return {'current_user': None}


FAMILY_ABBR = {
    "Access Control": "AC",
    "Awareness and Training": "AT",
    "Audit and Accountability": "AU",
    "Configuration Management": "CM",
    "Identification and Authentication": "IA",
    "Incident Response": "IR",
    "Maintenance": "MA",
    "Media Protection": "MP",
    "Personnel Security": "PS",
    "Physical Protection": "PE",
    "Risk Assessment": "RA",
    "Security Assessment": "CA",
    "System and Communications Protection": "SC",
    "System and Information Integrity": "SI",
}

FAMILY_COLORS = {
    "AC": "#3b82f6", "AT": "#8b5cf6", "AU": "#ec4899", "CM": "#f59e0b",
    "IA": "#10b981", "IR": "#ef4444", "MA": "#6366f1", "MP": "#14b8a6",
    "PS": "#f97316", "PE": "#84cc16", "RA": "#06b6d4", "CA": "#a855f7",
    "SC": "#0ea5e9", "SI": "#e11d48",
}


@app.route("/")
def dashboard():
    conn = get_db()
    families = conn.execute("""
        SELECT o.family, 
               COUNT(*) as total,
               SUM(CASE WHEN o.captured = 1 THEN 1 ELSE 0 END) as captured,
               SUM(CASE WHEN o.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
               SUM(CASE WHEN o.status = 'Evidence Collected' THEN 1 ELSE 0 END) as evidence_collected,
               SUM(CASE WHEN o.status = 'Reviewed' THEN 1 ELSE 0 END) as reviewed,
               COUNT(DISTINCT a.objective_id) as objectives_with_evidence,
               COALESCE(SUM(CASE WHEN a.id IS NOT NULL THEN 1 ELSE 0 END), 0) as artifact_count
        FROM objectives o
        LEFT JOIN artifacts a ON o.id = a.objective_id
        GROUP BY o.family ORDER BY o.family
    """).fetchall()
    totals = conn.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN o.captured = 1 THEN 1 ELSE 0 END) as captured,
               SUM(CASE WHEN o.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
               SUM(CASE WHEN o.status = 'Evidence Collected' THEN 1 ELSE 0 END) as evidence_collected,
               SUM(CASE WHEN o.status = 'Reviewed' THEN 1 ELSE 0 END) as reviewed,
               COUNT(DISTINCT a.objective_id) as objectives_with_evidence,
               COALESCE(SUM(CASE WHEN a.id IS NOT NULL THEN 1 ELSE 0 END), 0) as artifact_count
        FROM objectives o
        LEFT JOIN artifacts a ON o.id = a.objective_id
    """).fetchone()
    domain_coverage = conn.execute("""
        SELECT d.id, d.name, d.color, COUNT(DISTINCT a.objective_id) as obj_count
        FROM domains d
        LEFT JOIN artifacts a ON d.id = a.domain_id
        GROUP BY d.id ORDER BY d.name
    """).fetchall()
    # Count shared artifacts (linked to more than one objective)
    shared_artifacts = conn.execute("""
        SELECT COUNT(*) FROM (
            SELECT artifact_id FROM artifact_objectives
            GROUP BY artifact_id HAVING COUNT(*) > 1
        )
    """).fetchone()[0]
    conn.close()
    return render_template("dashboard.html",
                           families=families, totals=totals,
                           abbr=FAMILY_ABBR, colors=FAMILY_COLORS,
                           domain_coverage=domain_coverage,
                           shared_artifacts=shared_artifacts)


@app.route("/family/<path:family_name>")
def family_detail(family_name):
    conn = get_db()
    objectives = conn.execute("""
        SELECT * FROM objectives WHERE family = ? ORDER BY sort_as
    """, (family_name,)).fetchall()
    conn.close()

    # Group by requirement
    requirements = {}
    for obj in objectives:
        req_id = obj["requirement_id"]
        if req_id not in requirements:
            requirements[req_id] = {
                "id": req_id,
                "security_requirement": obj["security_requirement"],
                "examine": obj["examine"],
                "interview": obj["interview"],
                "test": obj["test"],
                "objectives": [],
            }
        requirements[req_id]["objectives"].append(dict(obj))

    abbr = FAMILY_ABBR.get(family_name, "")
    return render_template("family.html", family=family_name, abbr=abbr,
                           requirements=requirements,
                           color=FAMILY_COLORS.get(abbr, "#6366f1"))


@app.route("/api/toggle", methods=["POST"])
def toggle_objective():
    data = request.json
    obj_id = data.get("id")
    captured = data.get("captured", False)
    notes = data.get("artifact_notes", "")
    status = "Complete" if captured else "Not Started"
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d") if captured else None
    conn.execute("""
        UPDATE objectives SET captured = ?, artifact_notes = ?, captured_date = ?, status = ?
        WHERE id = ?
    """, (1 if captured else 0, notes, now, status, obj_id))
    log_audit('toggle', 'objective', obj_id,
              f"Set to {'Complete' if captured else 'Not Started'}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


VALID_STATUSES = ["Not Started", "In Progress", "Evidence Collected", "Reviewed", "Complete"]


@app.route("/api/status", methods=["POST"])
def update_status():
    data = request.json
    obj_id = data.get("id")
    status = data.get("status", "Not Started")
    if status not in VALID_STATUSES:
        return jsonify({"error": "Invalid status"}), 400
    conn = get_db()
    captured = 1 if status == "Complete" else 0
    now = datetime.now().strftime("%Y-%m-%d") if captured else None
    # Only set captured_date if transitioning to Complete
    if captured:
        existing = conn.execute("SELECT captured_date FROM objectives WHERE id = ?", (obj_id,)).fetchone()
        if existing and existing["captured_date"]:
            now = existing["captured_date"]  # Keep original date
    conn.execute("""
        UPDATE objectives SET status = ?, captured = ?, captured_date = COALESCE(?, captured_date)
        WHERE id = ?
    """, (status, captured, now, obj_id))
    log_audit('status_change', 'objective', obj_id, f"Set status to {status}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "status": status, "captured": captured})


@app.route("/api/notes", methods=["POST"])
def update_notes():
    data = request.json
    conn = get_db()
    obj_id = data.get("id")
    conn.execute("UPDATE objectives SET artifact_notes = ? WHERE id = ?",
                 (data.get("notes", ""), obj_id))
    log_audit('notes_saved', 'objective', obj_id, 'Notes updated', conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


### Comments / Activity Log ###

@app.route("/api/comments/<objective_id>")
def list_comments(objective_id):
    conn = get_db()
    rows = conn.execute("""
        SELECT * FROM objective_comments WHERE objective_id = ?
        ORDER BY created_at DESC
    """, (objective_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/comments", methods=["POST"])
def add_comment():
    data = request.json
    obj_id = data.get("objective_id")
    comment = data.get("comment", "").strip()
    if not obj_id or not comment:
        return jsonify({"error": "objective_id and comment required"}), 400
    username = session.get("username", "system")
    user_id = session.get("user_id")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    cur = conn.execute("""
        INSERT INTO objective_comments (objective_id, user_id, username, comment, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (obj_id, user_id, username, comment, now))
    comment_id = cur.lastrowid
    log_audit('comment_added', 'objective', obj_id, comment[:100], conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": comment_id, "username": username, "created_at": now})


@app.route("/api/comments/<int:comment_id>", methods=["DELETE"])
def delete_comment(comment_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM objective_comments WHERE id = ?", (comment_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found"}), 404
    # Only allow delete by comment author or admin
    if session.get("role") != "admin" and session.get("user_id") != row["user_id"]:
        conn.close()
        return jsonify({"error": "Not authorized"}), 403
    conn.execute("DELETE FROM objective_comments WHERE id = ?", (comment_id,))
    log_audit('comment_deleted', 'objective', row['objective_id'], f'Comment #{comment_id}', conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/comments/count")
def comment_counts():
    """Get comment counts for multiple objectives (used by family page)."""
    obj_ids = request.args.get("ids", "")
    if not obj_ids:
        return jsonify({})
    ids = [x.strip() for x in obj_ids.split(",") if x.strip()]
    conn = get_db()
    placeholders = ",".join("?" * len(ids))
    rows = conn.execute(f"""
        SELECT objective_id, COUNT(*) as cnt
        FROM objective_comments
        WHERE objective_id IN ({placeholders})
        GROUP BY objective_id
    """, ids).fetchall()
    conn.close()
    return jsonify({r["objective_id"]: r["cnt"] for r in rows})


@app.route("/api/bulk", methods=["POST"])
def bulk_update():
    data = request.json
    req_id = data.get("requirement_id")
    captured = data.get("captured", True)
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d") if captured else None
    status = "Complete" if captured else "Not Started"
    conn.execute("""
        UPDATE objectives SET captured = ?, captured_date = ?, status = ?
        WHERE requirement_id = ?
    """, (1 if captured else 0, now, status, req_id))
    log_audit('bulk_toggle', 'objective', req_id,
              f"Bulk set requirement {req_id} to {status}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/export")
def export_csv():
    conn = get_db()
    rows = conn.execute("SELECT * FROM objectives ORDER BY sort_as").fetchall()
    # Get all assignments with member names
    assignments = {}
    for a in conn.execute("""
        SELECT a.objective_id, t.name FROM artifact_assignments a
        JOIN team_members t ON a.member_id = t.id
        ORDER BY t.name
    """).fetchall():
        assignments.setdefault(a["objective_id"], []).append(a["name"])
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Family", "Requirement", "Objective", "Status",
                     "Captured", "Notes", "Date Captured", "Assigned To"])
    for r in rows:
        assigned = "; ".join(assignments.get(r["id"], []))
        writer.writerow([r["id"], r["family"], r["requirement_id"],
                         r["assessment_objective"],
                         r["status"] or "Not Started",
                         "Yes" if r["captured"] else "No",
                         r["artifact_notes"], r["captured_date"] or "",
                         assigned])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=cmmc-progress.csv"})


### POA&M (Plan of Action & Milestones) ###

@app.route("/poam")
def poam_page():
    conn = get_db()
    # Get all incomplete objectives with optional POA&M data
    rows = conn.execute("""
        SELECT o.id, o.family, o.requirement_id, o.assessment_objective, o.status,
               o.artifact_notes, o.security_requirement,
               p.weakness, p.remediation, p.resources, p.milestone_date, p.risk_level,
               GROUP_CONCAT(t.name, '; ') as assigned_to
        FROM objectives o
        LEFT JOIN poam p ON o.id = p.objective_id
        LEFT JOIN artifact_assignments a ON o.id = a.objective_id
        LEFT JOIN team_members t ON a.member_id = t.id
        WHERE o.captured = 0
        GROUP BY o.id
        ORDER BY o.sort_as
    """).fetchall()
    # Summary stats
    total_incomplete = len(rows)
    with_plan = sum(1 for r in rows if r['remediation'])
    by_risk = {'High': 0, 'Moderate': 0, 'Low': 0}
    for r in rows:
        risk = r['risk_level'] or 'Moderate'
        if risk in by_risk:
            by_risk[risk] += 1
    conn.close()
    return render_template("poam.html", items=rows, total=total_incomplete,
                           with_plan=with_plan, by_risk=by_risk,
                           abbr=FAMILY_ABBR, colors=FAMILY_COLORS)


@app.route("/api/poam", methods=["POST"])
def update_poam():
    data = request.json
    obj_id = data.get("objective_id")
    if not obj_id:
        return jsonify({"error": "objective_id required"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db()
    existing = conn.execute("SELECT id FROM poam WHERE objective_id = ?", (obj_id,)).fetchone()
    if existing:
        conn.execute("""
            UPDATE poam SET weakness=?, remediation=?, resources=?,
                   milestone_date=?, risk_level=?, updated_at=?
            WHERE objective_id=?
        """, (data.get("weakness", ""), data.get("remediation", ""),
              data.get("resources", ""), data.get("milestone_date") or None,
              data.get("risk_level", "Moderate"), now, obj_id))
    else:
        conn.execute("""
            INSERT INTO poam (objective_id, weakness, remediation, resources,
                              milestone_date, risk_level, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (obj_id, data.get("weakness", ""), data.get("remediation", ""),
              data.get("resources", ""), data.get("milestone_date") or None,
              data.get("risk_level", "Moderate"), now))
    log_audit('poam_saved', 'poam', obj_id,
              f"POA&M {'updated' if existing else 'created'} for {obj_id} (risk: {data.get('risk_level', 'Moderate')})",
              conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/poam/export")
def export_poam_csv():
    conn = get_db()
    rows = conn.execute("""
        SELECT o.id, o.family, o.requirement_id, o.assessment_objective, o.status,
               o.security_requirement,
               COALESCE(p.weakness, '') as weakness,
               COALESCE(p.remediation, '') as remediation,
               COALESCE(p.resources, '') as resources,
               COALESCE(p.milestone_date, '') as milestone_date,
               COALESCE(p.risk_level, 'Moderate') as risk_level,
               GROUP_CONCAT(t.name, '; ') as assigned_to
        FROM objectives o
        LEFT JOIN poam p ON o.id = p.objective_id
        LEFT JOIN artifact_assignments a ON o.id = a.objective_id
        LEFT JOIN team_members t ON a.member_id = t.id
        WHERE o.captured = 0
        GROUP BY o.id
        ORDER BY o.sort_as
    """).fetchall()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["POA&M ID", "Requirement", "Objective ID", "Assessment Objective",
                     "Status", "Risk Level", "Weakness", "Remediation Plan",
                     "Resources Required", "Milestone Date", "Responsible Party"])
    for i, r in enumerate(rows, 1):
        writer.writerow([f"POAM-{i:04d}", r["requirement_id"], r["id"],
                         r["assessment_objective"], r["status"] or "Not Started",
                         r["risk_level"], r["weakness"], r["remediation"],
                         r["resources"], r["milestone_date"],
                         r["assigned_to"] or "Unassigned"])
    return Response(output.getvalue(), mimetype="text/csv",
                    headers={"Content-Disposition": "attachment;filename=poam-report.csv"})


@app.route("/api/search")
def api_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    conn = get_db()
    results = conn.execute("""
        SELECT id, family, requirement_id, assessment_objective, captured, status
        FROM objectives
        WHERE id LIKE ? OR assessment_objective LIKE ? 
              OR family LIKE ? OR requirement_id LIKE ?
              OR security_requirement LIKE ?
        ORDER BY sort_as LIMIT 50
    """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
    conn.close()
    return jsonify([dict(r) for r in results])


@app.route("/search")
def search_page():
    q = request.args.get("q", "").strip()
    results = []
    if q:
        conn = get_db()
        results = conn.execute("""
            SELECT o.*, GROUP_CONCAT(t.name, '; ') as assigned_to
            FROM objectives o
            LEFT JOIN artifact_assignments a ON o.id = a.objective_id
            LEFT JOIN team_members t ON a.member_id = t.id
            WHERE o.id LIKE ? OR o.assessment_objective LIKE ? 
                  OR o.family LIKE ? OR o.requirement_id LIKE ?
                  OR o.security_requirement LIKE ?
            GROUP BY o.id
            ORDER BY o.sort_as
        """, (f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
        conn.close()
    return render_template("search.html", q=q, results=results, abbr=FAMILY_ABBR, colors=FAMILY_COLORS)


@app.route("/member/<int:member_id>")
def member_detail(member_id):
    conn = get_db()
    member = conn.execute("SELECT * FROM team_members WHERE id = ?", (member_id,)).fetchone()
    if not member:
        conn.close()
        return "Member not found", 404
    assignments = conn.execute("""
        SELECT a.*, o.family, o.assessment_objective, o.captured, o.captured_date, o.artifact_notes
        FROM artifact_assignments a
        JOIN objectives o ON a.objective_id = o.id
        WHERE a.member_id = ?
        ORDER BY o.sort_as
    """, (member_id,)).fetchall()
    # Stats
    total = len(assignments)
    completed = sum(1 for a in assignments if a["captured"])
    overdue = sum(1 for a in assignments if a["due_date"] and a["due_date"] < datetime.now().strftime("%Y-%m-%d") and not a["captured"])
    # Group by family
    by_family = {}
    for a in assignments:
        fam = a["family"]
        if fam not in by_family:
            by_family[fam] = {"total": 0, "completed": 0, "items": []}
        by_family[fam]["total"] += 1
        if a["captured"]:
            by_family[fam]["completed"] += 1
        by_family[fam]["items"].append(dict(a))
    conn.close()
    return render_template("member.html", member=member, assignments=assignments,
                           total=total, completed=completed, overdue=overdue,
                           by_family=by_family, abbr=FAMILY_ABBR, colors=FAMILY_COLORS,
                           now=datetime.now().strftime("%Y-%m-%d"))


### Config / Team Management ###

@app.route("/config")
@admin_required
def config_page():
    conn = get_db()
    members = conn.execute("SELECT * FROM team_members ORDER BY name").fetchall()
    # Get assignment counts per member
    counts = {}
    for m in members:
        row = conn.execute(
            "SELECT COUNT(*) as c FROM artifact_assignments WHERE member_id = ?",
            (m["id"],)
        ).fetchone()
        counts[m["id"]] = row["c"]
    domains = conn.execute("SELECT * FROM domains ORDER BY name").fetchall()
    conn.close()
    return render_template("config.html", members=members, counts=counts, domains=domains)


@app.route("/api/team", methods=["GET"])
def list_team():
    conn = get_db()
    members = conn.execute("SELECT * FROM team_members ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(m) for m in members])


@app.route("/api/team", methods=["POST"])
@admin_required
def add_team_member():
    data = request.json
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    conn = get_db()
    conn.execute(
        "INSERT INTO team_members (name, role, email, created_at) VALUES (?, ?, ?, ?)",
        (name, data.get("role", ""), data.get("email", ""),
         datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    log_audit('created', 'team', name, f"Created team member {name}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/team/<int:member_id>", methods=["DELETE"])
@admin_required
def delete_team_member(member_id):
    conn = get_db()
    member = conn.execute("SELECT name FROM team_members WHERE id = ?", (member_id,)).fetchone()
    conn.execute("DELETE FROM artifact_assignments WHERE member_id = ?", (member_id,))
    conn.execute("DELETE FROM team_members WHERE id = ?", (member_id,))
    log_audit('deleted', 'team', member_id,
              f"Deleted team member {member['name']}" if member else f"Deleted team member #{member_id}",
              conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/team/<int:member_id>", methods=["PATCH"])
@admin_required
def update_team_member(member_id):
    data = request.json
    conn = get_db()
    conn.execute(
        "UPDATE team_members SET name = ?, role = ?, email = ? WHERE id = ?",
        (data.get("name", ""), data.get("role", ""), data.get("email", ""), member_id)
    )
    log_audit('edited', 'team', member_id, f"Updated team member {data.get('name', '')}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


### Domains ###

@app.route("/api/domains", methods=["GET"])
def list_domains():
    conn = get_db()
    rows = conn.execute("SELECT * FROM domains ORDER BY name").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/domains", methods=["POST"])
@admin_required
def add_domain():
    data = request.json
    name = data.get("name", "").strip()
    color = data.get("color", "#6366f1").strip()
    if not name:
        return jsonify({"error": "Name required"}), 400
    conn = get_db()
    try:
        conn.execute("INSERT INTO domains (name, color) VALUES (?, ?)", (name, color))
        log_audit('created', 'domain', name, f"Created domain {name}", conn=conn)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Domain already exists"}), 409
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/domains/<int:domain_id>", methods=["PATCH"])
@admin_required
def update_domain(domain_id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE domains SET name = ?, color = ? WHERE id = ?",
                 (data.get("name", "").strip(), data.get("color", "#6366f1").strip(), domain_id))
    log_audit('edited', 'domain', domain_id, f"Updated domain {data.get('name', '')}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/domains/<int:domain_id>", methods=["DELETE"])
@admin_required
def delete_domain(domain_id):
    conn = get_db()
    domain = conn.execute("SELECT name FROM domains WHERE id = ?", (domain_id,)).fetchone()
    conn.execute("UPDATE artifacts SET domain_id = NULL WHERE domain_id = ?", (domain_id,))
    conn.execute("DELETE FROM domains WHERE id = ?", (domain_id,))
    log_audit('deleted', 'domain', domain_id,
              f"Deleted domain {domain['name']}" if domain else f"Deleted domain #{domain_id}",
              conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


### Assignments ###

@app.route("/api/assignments/<objective_id>")
def list_assignments(objective_id):
    conn = get_db()
    rows = conn.execute("""
        SELECT a.*, t.name as member_name, t.role as member_role
        FROM artifact_assignments a
        JOIN team_members t ON a.member_id = t.id
        WHERE a.objective_id = ?
        ORDER BY a.assigned_at DESC
    """, (objective_id,)).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/assignments", methods=["POST"])
def add_assignment():
    data = request.json
    objective_id = data.get("objective_id")
    member_id = data.get("member_id")
    if not objective_id or not member_id:
        return jsonify({"error": "objective_id and member_id required"}), 400
    conn = get_db()
    # Prevent duplicate assignments
    existing = conn.execute(
        "SELECT id FROM artifact_assignments WHERE objective_id = ? AND member_id = ?",
        (objective_id, member_id)
    ).fetchone()
    if existing:
        conn.close()
        return jsonify({"error": "Already assigned"}), 409
    conn.execute("""
        INSERT INTO artifact_assignments (objective_id, member_id, status, due_date, assigned_at)
        VALUES (?, ?, ?, ?, ?)
    """, (objective_id, member_id, data.get("status", "assigned"),
          data.get("due_date"), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    member = conn.execute("SELECT name FROM team_members WHERE id = ?", (member_id,)).fetchone()
    member_name = member['name'] if member else f"#{member_id}"
    log_audit('created', 'assignment', objective_id,
              f"Assigned {objective_id} to {member_name}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/assignments/<int:assignment_id>", methods=["DELETE"])
def delete_assignment(assignment_id):
    conn = get_db()
    asgn = conn.execute("""
        SELECT a.objective_id, t.name FROM artifact_assignments a
        JOIN team_members t ON a.member_id = t.id WHERE a.id = ?
    """, (assignment_id,)).fetchone()
    conn.execute("DELETE FROM artifact_assignments WHERE id = ?", (assignment_id,))
    detail = f"Removed assignment {asgn['objective_id']} from {asgn['name']}" if asgn else f"Deleted assignment #{assignment_id}"
    log_audit('deleted', 'assignment', assignment_id, detail, conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/assignments/<int:assignment_id>/due", methods=["PATCH"])
def update_assignment_due(assignment_id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE artifact_assignments SET due_date = ? WHERE id = ?",
                 (data.get("due_date") or None, assignment_id))
    log_audit('due_date_changed', 'assignment', assignment_id,
              f"Due date set to {data.get('due_date') or 'none'}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/assignments/<int:assignment_id>/status", methods=["PATCH"])
def update_assignment_status(assignment_id):
    data = request.json
    conn = get_db()
    conn.execute("UPDATE artifact_assignments SET status = ? WHERE id = ?",
                 (data.get("status", "assigned"), assignment_id))
    log_audit('status_change', 'assignment', assignment_id,
              f"Assignment status set to {data.get('status', 'assigned')}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/assignments/bulk", methods=["POST"])
def bulk_assign():
    data = request.json
    requirement_id = data.get("requirement_id")
    member_id = data.get("member_id")
    if not requirement_id or not member_id:
        return jsonify({"error": "requirement_id and member_id required"}), 400
    conn = get_db()
    objectives = conn.execute(
        "SELECT id FROM objectives WHERE requirement_id = ?", (requirement_id,)
    ).fetchall()
    added = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for obj in objectives:
        existing = conn.execute(
            "SELECT id FROM artifact_assignments WHERE objective_id = ? AND member_id = ?",
            (obj["id"], member_id)
        ).fetchone()
        if not existing:
            conn.execute("""
                INSERT INTO artifact_assignments (objective_id, member_id, status, assigned_at)
                VALUES (?, ?, 'assigned', ?)
            """, (obj["id"], member_id, now))
            added += 1
    member = conn.execute("SELECT name FROM team_members WHERE id = ?", (member_id,)).fetchone()
    member_name = member['name'] if member else f"#{member_id}"
    log_audit('bulk_assign', 'assignment', requirement_id,
              f"Bulk assigned {requirement_id} to {member_name} ({added} new of {len(objectives)} objectives)",
              conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "added": added, "total": len(objectives)})


### Artifacts ###

@app.route("/api/artifacts/<objective_id>")
def list_artifacts(objective_id):
    conn = get_db()
    rows = conn.execute("""
        SELECT a.*, d.name as domain_name, d.color as domain_color,
               CASE WHEN a.objective_id != ? THEN 1 ELSE 0 END as is_linked
        FROM artifact_objectives ao
        JOIN artifacts a ON ao.artifact_id = a.id
        LEFT JOIN domains d ON a.domain_id = d.id
        WHERE ao.objective_id = ? ORDER BY a.uploaded_at DESC
    """, (objective_id, objective_id)).fetchall()
    # For each artifact, get all linked objectives
    result = []
    for r in rows:
        d = dict(r)
        linked = conn.execute("""
            SELECT ao.objective_id, o.family FROM artifact_objectives ao
            JOIN objectives o ON ao.objective_id = o.id
            WHERE ao.artifact_id = ?
        """, (r['id'],)).fetchall()
        d['linked_objectives'] = [{'id': l['objective_id'], 'family': l['family']} for l in linked]
        result.append(d)
    conn.close()
    return jsonify(result)


def _extract_file_created(filepath, ext):
    """Extract creation/modification date from file metadata."""
    try:
        # Images (EXIF)
        if ext in ('.jpg', '.jpeg', '.png', '.gif', '.webp'):
            from PIL import Image
            from PIL.ExifTags import Base as ExifBase
            img = Image.open(filepath)
            exif = img.getexif()
            if exif:
                # Try DateTimeOriginal, then DateTimeDigitized, then DateTime
                for tag in (ExifBase.DateTimeOriginal, ExifBase.DateTimeDigitized, ExifBase.DateTime):
                    val = exif.get(tag)
                    if val:
                        # EXIF format: "2026:03:05 10:30:00"
                        return val.replace(":", "-", 2)
            img.close()

        # PDFs
        elif ext == '.pdf':
            from pypdf import PdfReader
            reader = PdfReader(filepath)
            info = reader.metadata
            if info:
                for field in (info.creation_date, info.modification_date):
                    if field:
                        return field.strftime("%Y-%m-%d %H:%M:%S")

        # Word docs (.docx)
        elif ext == '.docx':
            from docx import Document
            doc = Document(filepath)
            props = doc.core_properties
            if props.created:
                return props.created.strftime("%Y-%m-%d %H:%M:%S")
            if props.modified:
                return props.modified.strftime("%Y-%m-%d %H:%M:%S")

        # Excel (.xlsx)
        elif ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            props = wb.properties
            if props.created:
                return props.created.strftime("%Y-%m-%d %H:%M:%S")
            if props.modified:
                return props.modified.strftime("%Y-%m-%d %H:%M:%S")
            wb.close()

        # PowerPoint (.pptx) - uses same OPC format as docx
        elif ext == '.pptx':
            import zipfile
            from xml.etree import ElementTree
            with zipfile.ZipFile(filepath) as z:
                if 'docProps/core.xml' in z.namelist():
                    tree = ElementTree.parse(z.open('docProps/core.xml'))
                    ns = {'dcterms': 'http://purl.org/dc/terms/'}
                    created = tree.find('.//dcterms:created', ns)
                    if created is not None and created.text:
                        return created.text[:19].replace("T", " ")

    except Exception:
        pass
    return None


def _generate_artifact_filename(conn, objective_id, domain_name, ext):
    """Generate auto-renamed filename: AC-3.01.01.a-Domain{ext}"""
    # Clean objective_id: 3.1.1[a] -> 3.01.01.a
    clean_id = objective_id.replace("[", ".").replace("]", "").replace(" ", "").replace("\t", "")
    # Get family abbreviation prefix from the objective
    obj_row = conn.execute("SELECT family FROM objectives WHERE id = ?", (objective_id,)).fetchone()
    abbr = FAMILY_ABBR.get(obj_row["family"], "") if obj_row else ""
    prefix = f"{abbr}-{clean_id}" if abbr else clean_id

    # Domain part
    domain_part = ""
    if domain_name:
        domain_part = "-" + domain_name.replace(" ", "-")

    return f"{prefix}{domain_part}{ext}"


@app.route("/api/upload", methods=["POST"])
def upload_artifact():
    objective_id = request.form.get("objective_id")
    if not objective_id:
        return jsonify({"error": "Missing objective_id"}), 400

    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "No file provided"}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return jsonify({"error": f"File type {ext} not allowed"}), 400

    domain_id = request.form.get("domain_id") or None
    if domain_id:
        domain_id = int(domain_id)

    # Create per-objective subdirectory
    safe_id = objective_id.replace("[", "").replace("]", "").replace(" ", "").replace("\t", "")
    obj_dir = os.path.join(UPLOAD_DIR, safe_id)
    os.makedirs(obj_dir, exist_ok=True)

    conn = get_db()

    # Look up domain name for filename
    domain_name = ""
    if domain_id:
        d = conn.execute("SELECT name FROM domains WHERE id = ?", (domain_id,)).fetchone()
        if d:
            domain_name = d["name"]

    # Auto-rename
    filename = _generate_artifact_filename(conn, objective_id, domain_name, ext)
    filepath = os.path.join(obj_dir, filename)

    # Avoid collision
    if os.path.exists(filepath):
        timestamp = datetime.now().strftime("%H%M%S")
        filename = _generate_artifact_filename(conn, objective_id, domain_name, f"_{timestamp}{ext}")
        filepath = os.path.join(obj_dir, filename)

    file.save(filepath)
    file_size = os.path.getsize(filepath)

    # Extract creation date from file metadata
    file_created = _extract_file_created(filepath, ext)

    now_ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    cursor = conn.execute("""
        INSERT INTO artifacts (objective_id, filename, original_name, file_size, mime_type, uploaded_at, domain_id, file_created)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (objective_id, f"{safe_id}/{filename}", file.filename, file_size,
          file.content_type, now_ts, domain_id, file_created))
    # Also insert into junction table
    conn.execute("INSERT OR IGNORE INTO artifact_objectives (artifact_id, objective_id) VALUES (?, ?)",
                  (cursor.lastrowid, objective_id))
    log_audit('uploaded', 'artifact', objective_id, f"Uploaded {filename}", conn=conn)
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "filename": filename})


@app.route("/api/artifacts/delete/<int:artifact_id>", methods=["POST"])
def delete_artifact(artifact_id):
    conn = get_db()
    row = conn.execute("SELECT * FROM artifacts WHERE id = ?", (artifact_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    # Delete file
    filepath = os.path.join(UPLOAD_DIR, row["filename"])
    if os.path.exists(filepath):
        os.remove(filepath)

    conn.execute("DELETE FROM artifact_objectives WHERE artifact_id = ?", (artifact_id,))
    conn.execute("DELETE FROM artifacts WHERE id = ?", (artifact_id,))
    log_audit('deleted', 'artifact', artifact_id,
              f"Deleted {row['original_name']} from {row['objective_id']}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/artifacts/<int:artifact_id>/domain", methods=["PATCH"])
def update_artifact_domain(artifact_id):
    data = request.json
    domain_id = data.get("domain_id") or None
    if domain_id:
        domain_id = int(domain_id)

    conn = get_db()
    row = conn.execute("SELECT * FROM artifacts WHERE id = ?", (artifact_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    old_filepath = os.path.join(UPLOAD_DIR, row["filename"])
    ext = os.path.splitext(row["filename"])[1]

    # Look up domain name
    domain_name = ""
    if domain_id:
        d = conn.execute("SELECT name FROM domains WHERE id = ?", (domain_id,)).fetchone()
        if d:
            domain_name = d["name"]

    # Generate new filename
    new_basename = _generate_artifact_filename(conn, row["objective_id"], domain_name, ext)
    safe_id = row["objective_id"].replace("[", "").replace("]", "").replace(" ", "").replace("\t", "")
    new_filename = f"{safe_id}/{new_basename}"
    new_filepath = os.path.join(UPLOAD_DIR, new_filename)

    # Rename physical file
    if os.path.exists(old_filepath):
        os.makedirs(os.path.dirname(new_filepath), exist_ok=True)
        os.rename(old_filepath, new_filepath)

    # Update DB
    conn.execute("UPDATE artifacts SET domain_id = ?, filename = ? WHERE id = ?",
                 (domain_id, new_filename, artifact_id))
    log_audit('domain_changed', 'artifact', artifact_id,
              f"Domain set to {domain_name or 'none'} for {row['objective_id']}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "filename": new_filename})


@app.route("/api/artifacts/<int:artifact_id>/obtained", methods=["PATCH"])
def update_artifact_obtained(artifact_id):
    data = request.json
    method = data.get("obtained_method", "")
    conn = get_db()
    conn.execute("UPDATE artifacts SET obtained_method = ? WHERE id = ?", (method, artifact_id))
    log_audit('obtained_updated', 'artifact', artifact_id,
              f"Obtained method set to '{method}'" if method else "Obtained method cleared", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/artifacts/<int:artifact_id>/link", methods=["POST"])
def link_artifact(artifact_id):
    data = request.json
    objective_id = data.get("objective_id", "").strip()
    if not objective_id:
        return jsonify({"error": "Missing objective_id"}), 400
    conn = get_db()
    # Verify artifact and objective exist
    artifact = conn.execute("SELECT id FROM artifacts WHERE id = ?", (artifact_id,)).fetchone()
    if not artifact:
        conn.close()
        return jsonify({"error": "Artifact not found"}), 404
    objective = conn.execute("SELECT id FROM objectives WHERE id = ?", (objective_id,)).fetchone()
    if not objective:
        conn.close()
        return jsonify({"error": "Objective not found"}), 404
    try:
        conn.execute("INSERT INTO artifact_objectives (artifact_id, objective_id) VALUES (?, ?)",
                      (artifact_id, objective_id))
        log_audit('linked', 'artifact', artifact_id,
                  f"Linked artifact #{artifact_id} to {objective_id}", conn=conn)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Already linked"}), 409
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/artifacts/<int:artifact_id>/link/<path:objective_id>", methods=["DELETE"])
def unlink_artifact(artifact_id, objective_id):
    conn = get_db()
    # Don't allow unlinking the original/primary objective if it's the last link
    count = conn.execute("SELECT COUNT(*) FROM artifact_objectives WHERE artifact_id = ?",
                          (artifact_id,)).fetchone()[0]
    if count <= 1:
        conn.close()
        return jsonify({"error": "Cannot unlink the last objective. Delete the artifact instead."}), 400
    conn.execute("DELETE FROM artifact_objectives WHERE artifact_id = ? AND objective_id = ?",
                  (artifact_id, objective_id))
    log_audit('unlinked', 'artifact', artifact_id,
              f"Unlinked artifact #{artifact_id} from {objective_id}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/artifacts/library")
def artifacts_library_api():
    conn = get_db()
    rows = conn.execute("""
        SELECT a.*, d.name as domain_name, d.color as domain_color
        FROM artifacts a
        LEFT JOIN domains d ON a.domain_id = d.id
        ORDER BY a.uploaded_at DESC
    """).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        linked = conn.execute("""
            SELECT ao.objective_id, o.family FROM artifact_objectives ao
            JOIN objectives o ON ao.objective_id = o.id
            WHERE ao.artifact_id = ?
        """, (r['id'],)).fetchall()
        d['linked_objectives'] = [{'id': l['objective_id'], 'family': l['family']} for l in linked]
        result.append(d)
    conn.close()
    return jsonify(result)


@app.route("/artifacts")
def artifacts_library_page():
    return render_template("artifacts.html")


@app.route("/uploads/<path:filename>")
def serve_upload(filename):
    from flask import send_from_directory
    # Resolve the real path and verify it's within UPLOAD_DIR
    real_upload = os.path.realpath(UPLOAD_DIR)
    real_file = os.path.realpath(os.path.join(UPLOAD_DIR, filename))
    if not real_file.startswith(real_upload + os.sep) and real_file != real_upload:
        abort(404)
    return send_from_directory(UPLOAD_DIR, filename)


@app.route("/api/hash-artifacts", methods=["POST"])
def hash_artifacts():
    """Generate CMMC-compliant SHA-256 hashes for all uploaded artifacts."""
    import hashlib
    results = []
    for root, dirs, files in os.walk(UPLOAD_DIR):
        for fname in sorted(files):
            fpath = os.path.join(root, fname)
            sha256 = hashlib.sha256()
            with open(fpath, "rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    sha256.update(chunk)
            rel_path = os.path.relpath(fpath, UPLOAD_DIR)
            results.append({
                "algorithm": "SHA256",
                "hash": sha256.hexdigest().upper(),
                "path": rel_path
            })

    # Write CMMCAssessmentArtifacts.log
    log_lines = [f"{'Algorithm':<12} {'Hash':<64} Path"]
    log_lines.append(f"{'-'*12} {'-'*64} {'-'*4}")
    for r in results:
        log_lines.append(f"{r['algorithm']:<12} {r['hash']:<64} {r['path']}")
    log_content = "\n".join(log_lines)

    artifacts_log = os.path.join(UPLOAD_DIR, "CMMCAssessmentArtifacts.log")
    with open(artifacts_log, "w", encoding="ascii") as f:
        f.write(log_content)

    # Hash the log file itself
    sha256 = hashlib.sha256()
    with open(artifacts_log, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            sha256.update(chunk)
    log_hash = sha256.hexdigest().upper()

    hash_log = os.path.join(UPLOAD_DIR, "CMMCAssessmentLogHash.log")
    with open(hash_log, "w", encoding="ascii") as f:
        f.write(f"{'Algorithm':<12} {'Hash':<64} Path\n")
        f.write(f"{'-'*12} {'-'*64} {'-'*4}\n")
        f.write(f"{'SHA256':<12} {log_hash:<64} CMMCAssessmentArtifacts.log\n")

    log_audit('hash_generated', 'artifact', None,
              f"Generated SHA-256 hashes for {len(results)} artifacts")

    return jsonify({
        "ok": True,
        "artifacts_hashed": len(results),
        "log_hash": log_hash,
        "artifacts_log": "uploads/CMMCAssessmentArtifacts.log",
        "hash_log": "uploads/CMMCAssessmentLogHash.log"
    })


### Authentication ###

@app.route("/setup", methods=["GET", "POST"])
def setup():
    conn = get_db()
    user_count = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    conn.close()
    if user_count > 0:
        return redirect(url_for('login'))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        first_name = request.form.get("first_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        confirm = request.form.get("confirm_password", "")
        errors = []
        if not username:
            errors.append("Username is required")
        if not first_name or not last_name:
            errors.append("First and last name are required")
        if password != confirm:
            errors.append("Passwords do not match")
        errors.extend(validate_password(password))
        if errors:
            return render_template("setup.html", errors=errors, username=username, first_name=first_name, last_name=last_name)
        conn = get_db()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        display_name = f"{first_name} {last_name}"
        conn.execute(
            "INSERT INTO users (username, password_hash, role, first_name, last_name, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (username, generate_password_hash(password), 'admin', first_name, last_name, now)
        )
        # Auto-create team member for assignment dropdown
        conn.execute(
            "INSERT INTO team_members (name, role, email, created_at) VALUES (?, ?, ?, ?)",
            (display_name, 'admin', '', now)
        )
        conn.commit()
        # Log after commit since user now exists — use manual insert since session isn't set yet
        conn2 = get_db()
        user_row = conn2.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
        conn2.execute(
            "INSERT INTO audit_log (user_id, username, action, target_type, target_id, details, timestamp) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (user_row['id'], username, 'created', 'user', str(user_row['id']),
             f"Initial admin user {first_name} {last_name} ({username}) created via setup",
             datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn2.commit()
        conn2.close()
        conn.close()
        return redirect(url_for('login'))
    return render_template("setup.html", errors=[], username="")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        # Rate limiting
        client_ip = request.remote_addr
        if not _check_rate_limit(client_ip):
            return render_template("login.html", error="Too many login attempts. Please wait 5 minutes.")
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        if user and check_password_hash(user['password_hash'], password):
            session.clear()  # Prevent session fixation
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            return redirect(url_for('dashboard'))
        return render_template("login.html", error="Invalid username or password")
    return render_template("login.html", error=None)


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route("/admin/users")
@admin_required
def admin_users():
    conn = get_db()
    users = conn.execute("SELECT id, username, role, first_name, last_name, created_at FROM users ORDER BY created_at").fetchall()
    conn.close()
    return render_template("admin_users.html", users=users)


@app.route("/api/admin/users", methods=["POST"])
@admin_required
def admin_create_user():
    data = request.json
    username = data.get("username", "").strip()
    password = data.get("password", "")
    role = data.get("role", "user")
    first_name = data.get("first_name", "").strip()
    last_name = data.get("last_name", "").strip()
    if role not in ('admin', 'user'):
        return jsonify({"error": "Invalid role"}), 400
    if not username:
        return jsonify({"error": "Username is required"}), 400
    if not first_name or not last_name:
        return jsonify({"error": "First and last name are required"}), 400
    errors = validate_password(password)
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    conn = get_db()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    display_name = f"{first_name} {last_name}"
    try:
        conn.execute(
            "INSERT INTO users (username, password_hash, role, first_name, last_name, created_at) VALUES (?, ?, ?, ?, ?, ?)",
            (username, generate_password_hash(password), role, first_name, last_name, now)
        )
        # Auto-create team member for assignment dropdown
        conn.execute(
            "INSERT INTO team_members (name, role, email, created_at) VALUES (?, ?, ?, ?)",
            (display_name, role, '', now)
        )
        log_audit('created', 'user', username,
                  f"Created user {first_name} {last_name} ({username}) as {role}", conn=conn)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Username already exists"}), 409
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:user_id>", methods=["PATCH"])
@admin_required
def admin_edit_user(user_id):
    data = request.json
    first_name = data.get("first_name", "").strip()
    last_name = data.get("last_name", "").strip()
    role = data.get("role", "").strip()
    if not first_name or not last_name:
        return jsonify({"error": "First and last name are required"}), 400
    if role not in ('admin', 'user'):
        return jsonify({"error": "Invalid role"}), 400
    conn = get_db()
    old_user = conn.execute("SELECT username, first_name, last_name, role FROM users WHERE id = ?", (user_id,)).fetchone()
    if not old_user:
        conn.close()
        return jsonify({"error": "User not found"}), 404
    # Prevent removing own admin role
    if user_id == session.get('user_id') and role != 'admin':
        conn.close()
        return jsonify({"error": "Cannot remove your own admin role"}), 400
    old_display = f"{old_user['first_name']} {old_user['last_name']}"
    new_display = f"{first_name} {last_name}"
    conn.execute("UPDATE users SET first_name = ?, last_name = ?, role = ? WHERE id = ?",
                 (first_name, last_name, role, user_id))
    # Update team member name if it matches
    if old_display != new_display:
        conn.execute("UPDATE team_members SET name = ? WHERE name = ?", (new_display, old_display))
    if old_user['role'] != role:
        conn.execute("UPDATE team_members SET role = ? WHERE name = ?", (role, new_display))
    conn.commit()
    # Audit
    changes = []
    if old_user['first_name'] != first_name or old_user['last_name'] != last_name:
        changes.append(f"Name: {old_display} -> {new_display}")
    if old_user['role'] != role:
        changes.append(f"Role: {old_user['role']} -> {role}")
    if changes:
        log_audit('edit_user', 'user', str(user_id), '; '.join(changes), conn=conn)
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
@admin_required
def admin_delete_user(user_id):
    if user_id == session.get('user_id'):
        return jsonify({"error": "Cannot delete yourself"}), 400
    conn = get_db()
    user = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    # Also remove from team members
    if user:
        conn.execute("DELETE FROM artifact_assignments WHERE member_id IN (SELECT id FROM team_members WHERE name = ?)", (user['username'],))
        conn.execute("DELETE FROM team_members WHERE name = ?", (user['username'],))
    log_audit('deleted', 'user', user_id,
              f"Deleted user {user['username']}" if user else f"Deleted user #{user_id}", conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/admin/users/<int:user_id>/reset", methods=["POST"])
@admin_required
def admin_reset_password(user_id):
    data = request.json
    password = data.get("password", "")
    errors = validate_password(password)
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400
    conn = get_db()
    user = conn.execute("SELECT username FROM users WHERE id = ?", (user_id,)).fetchone()
    conn.execute("UPDATE users SET password_hash = ? WHERE id = ?",
                 (generate_password_hash(password), user_id))
    log_audit('password_reset', 'user', user_id,
              f"Password reset for {user['username']}" if user else f"Password reset for user #{user_id}",
              conn=conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/audit")
@admin_required
def audit_log_page():
    page = request.args.get("page", 1, type=int)
    per_page = 50
    filter_user = request.args.get("user", "")
    filter_action = request.args.get("action", "")
    filter_from = request.args.get("from", "")
    filter_to = request.args.get("to", "")

    conn = get_db()
    conditions = []
    params = []
    if filter_user:
        conditions.append("username = ?")
        params.append(filter_user)
    if filter_action:
        conditions.append("action = ?")
        params.append(filter_action)
    if filter_from:
        conditions.append("timestamp >= ?")
        params.append(filter_from + " 00:00:00")
    if filter_to:
        conditions.append("timestamp <= ?")
        params.append(filter_to + " 23:59:59")

    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    total = conn.execute(f"SELECT COUNT(*) FROM audit_log {where}", params).fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = min(page, total_pages)
    offset = (page - 1) * per_page

    rows = conn.execute(
        f"SELECT * FROM audit_log {where} ORDER BY timestamp DESC LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()
    users = conn.execute("SELECT DISTINCT username FROM audit_log WHERE username IS NOT NULL ORDER BY username").fetchall()
    actions = conn.execute("SELECT DISTINCT action FROM audit_log ORDER BY action").fetchall()
    conn.close()

    return render_template("audit.html", logs=rows, page=page, total_pages=total_pages,
                           total=total, users=[u[0] for u in users],
                           actions=[a[0] for a in actions],
                           filter_user=filter_user, filter_action=filter_action,
                           filter_from=filter_from, filter_to=filter_to)


@app.route("/report")
def report():
    conn = get_db()
    # Overall totals
    totals = conn.execute("""
        SELECT COUNT(*) as total,
               SUM(CASE WHEN o.captured = 1 THEN 1 ELSE 0 END) as captured,
               SUM(CASE WHEN o.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
               SUM(CASE WHEN o.status = 'Evidence Collected' THEN 1 ELSE 0 END) as evidence_collected,
               SUM(CASE WHEN o.status = 'Reviewed' THEN 1 ELSE 0 END) as reviewed
        FROM objectives o
    """).fetchone()
    totals = dict(totals)

    # Per-family summary (for exec summary table)
    families = conn.execute("""
        SELECT o.family,
               COUNT(*) as total,
               SUM(CASE WHEN o.captured = 1 THEN 1 ELSE 0 END) as captured,
               SUM(CASE WHEN o.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
               SUM(CASE WHEN o.status = 'Evidence Collected' THEN 1 ELSE 0 END) as evidence_collected,
               SUM(CASE WHEN o.status = 'Reviewed' THEN 1 ELSE 0 END) as reviewed,
               COALESCE(SUM(CASE WHEN a.id IS NOT NULL THEN 1 ELSE 0 END), 0) as artifact_count
        FROM objectives o
        LEFT JOIN artifacts a ON o.id = a.objective_id
        GROUP BY o.family ORDER BY o.family
    """).fetchall()
    families = [dict(f) for f in families]

    # All objectives with artifacts and assignments, grouped by family then requirement
    all_objectives = conn.execute("""
        SELECT * FROM objectives ORDER BY sort_as
    """).fetchall()

    # Prefetch all artifacts via artifact_objectives junction
    all_artifacts = conn.execute("""
        SELECT ao.objective_id, a.id, a.original_name, a.obtained_method,
               d.name as domain_name
        FROM artifact_objectives ao
        JOIN artifacts a ON ao.artifact_id = a.id
        LEFT JOIN domains d ON a.domain_id = d.id
        ORDER BY a.uploaded_at DESC
    """).fetchall()

    # Prefetch all assignments
    all_assignments = conn.execute("""
        SELECT aa.objective_id, t.name, aa.status
        FROM artifact_assignments aa
        JOIN team_members t ON aa.member_id = t.id
    """).fetchall()
    conn.close()

    # Index artifacts and assignments by objective_id
    artifacts_by_obj = {}
    for art in all_artifacts:
        artifacts_by_obj.setdefault(art["objective_id"], []).append(dict(art))
    assignments_by_obj = {}
    for asn in all_assignments:
        assignments_by_obj.setdefault(asn["objective_id"], []).append(dict(asn))

    # Build family_details: {family_name: {total, captured, artifact_count, requirements: {req_id: {security_requirement, discussion, objectives: [...]}}}}
    family_details = {}
    for obj in all_objectives:
        obj = dict(obj)
        fname = obj["family"]
        if fname not in family_details:
            family_details[fname] = {"total": 0, "captured": 0, "artifact_count": 0, "requirements": {}}
        fd = family_details[fname]
        fd["total"] += 1
        if obj.get("captured"):
            fd["captured"] += 1

        req_id = obj["requirement_id"]
        if req_id not in fd["requirements"]:
            fd["requirements"][req_id] = {
                "security_requirement": obj.get("security_requirement", ""),
                "discussion": obj.get("discussion", ""),
                "objectives": []
            }

        obj["artifacts"] = artifacts_by_obj.get(obj["id"], [])
        obj["assignments"] = assignments_by_obj.get(obj["id"], [])
        fd["artifact_count"] += len(obj["artifacts"])
        fd["requirements"][req_id]["objectives"].append(obj)

    pct_complete = round(totals["captured"] / totals["total"] * 100, 1) if totals["total"] else 0
    org_name = os.environ.get("ORG_NAME", "Organization")

    generated_date = datetime.now().strftime("%B %d, %Y")

    return render_template("report.html",
                           totals=totals, families=families,
                           family_details=family_details,
                           pct_complete=pct_complete,
                           org_name=org_name,
                           generated_date=generated_date,
                           abbr=FAMILY_ABBR, colors=FAMILY_COLORS)


@app.route("/landing")
def landing():
    return render_template("landing.html")


if __name__ == "__main__":
    init_db()
    # Seed example artifacts if not already done
    from seed_examples import seed_examples
    from seed_examples_supplement import seed_supplement
    conn_check = get_db()
    try:
        sample = conn_check.execute("SELECT example_artifacts FROM objectives LIMIT 1").fetchone()
        if sample and not sample["example_artifacts"]:
            conn_check.close()
            seed_examples()
            seed_supplement()
        else:
            conn_check.close()
    except Exception:
        conn_check.close()
        try:
            conn_check2 = get_db()
            conn_check2.execute("ALTER TABLE objectives ADD COLUMN example_artifacts TEXT DEFAULT ''")
            conn_check2.commit()
            conn_check2.close()
            seed_examples()
            seed_supplement()
        except Exception:
            pass

    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=3300, debug=debug)
