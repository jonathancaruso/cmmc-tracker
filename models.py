"""Database initialization, migrations, and SQL helper functions."""

import os
import re
import sqlite3

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get("DB_PATH", os.path.join(BASE_DIR, "cmmc.db"))
XLSX_PATH = os.path.join(BASE_DIR, "nist-800-171a.xlsx")
XLSX_171_PATH = os.path.join(BASE_DIR, "nist-800-171.xlsx")
UPLOAD_DIR = os.environ.get("UPLOAD_PATH", os.path.join(BASE_DIR, "uploads"))
os.makedirs(UPLOAD_DIR, exist_ok=True)

FAMILY_ABBR = {
    "Access Control": "AC",
    "Awareness and Training": "AT",
    "Audit and Accountability": "AU",
    "Configuration Management": "CM",
    "Identification and Authentication": "IA",
    "Incident Response": "IR",
    "Maintenance": "MA",
    "Media Protection": "MP",
    "Personnel Security": "PS",
    "Physical Protection": "PE",
    "Risk Assessment": "RA",
    "Security Assessment": "CA",
    "System and Communications Protection": "SC",
    "System and Information Integrity": "SI",
}

FAMILY_COLORS = {
    "AC": "#3b82f6", "AT": "#8b5cf6", "AU": "#ec4899", "CM": "#f59e0b",
    "IA": "#10b981", "IR": "#ef4444", "MA": "#6366f1", "MP": "#14b8a6",
    "PS": "#f97316", "PE": "#84cc16", "RA": "#06b6d4", "CA": "#a855f7",
    "SC": "#0ea5e9", "SI": "#e11d48",
}

VALID_STATUSES = ["Not Started", "In Progress", "Evidence Collected", "Reviewed", "Complete"]

ALLOWED_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.webp', '.pdf', '.doc', '.docx',
                      '.xls', '.xlsx', '.txt', '.csv', '.pptx', '.zip', '.md'}


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS objectives (
            id TEXT PRIMARY KEY,
            family TEXT NOT NULL,
            requirement_id TEXT NOT NULL,
            sort_as TEXT,
            security_requirement TEXT,
            assessment_objective TEXT,
            examine TEXT,
            interview TEXT,
            test TEXT,
            captured INTEGER DEFAULT 0,
            artifact_notes TEXT DEFAULT '',
            captured_date TEXT,
            requirement_type TEXT DEFAULT '',
            discussion TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS artifacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL,
            filename TEXT NOT NULL,
            original_name TEXT NOT NULL,
            file_size INTEGER,
            mime_type TEXT,
            uploaded_at TEXT NOT NULL,
            file_created TEXT,
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS team_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            role TEXT DEFAULT '',
            email TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS domains (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            color TEXT NOT NULL DEFAULT '#6366f1'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS artifact_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL,
            member_id INTEGER NOT NULL,
            status TEXT DEFAULT 'assigned',
            due_date TEXT,
            assigned_at TEXT NOT NULL,
            FOREIGN KEY (objective_id) REFERENCES objectives(id),
            FOREIGN KEY (member_id) REFERENCES team_members(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'user',
            first_name TEXT DEFAULT '',
            last_name TEXT DEFAULT '',
            created_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS poam (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL UNIQUE,
            weakness TEXT DEFAULT '',
            remediation TEXT DEFAULT '',
            resources TEXT DEFAULT '',
            milestone_date TEXT,
            risk_level TEXT DEFAULT 'Moderate',
            updated_at TEXT NOT NULL,
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            action TEXT NOT NULL,
            target_type TEXT,
            target_id TEXT,
            details TEXT,
            timestamp TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS objective_comments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            objective_id TEXT NOT NULL,
            user_id INTEGER,
            username TEXT NOT NULL,
            comment TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS artifact_objectives (
            artifact_id INTEGER NOT NULL,
            objective_id TEXT NOT NULL,
            PRIMARY KEY (artifact_id, objective_id),
            FOREIGN KEY (artifact_id) REFERENCES artifacts(id),
            FOREIGN KEY (objective_id) REFERENCES objectives(id)
        )
    """)
    conn.commit()

    # Migrate existing DBs: add columns if they don't exist
    for col in ["requirement_type", "discussion"]:
        try:
            conn.execute(f"SELECT {col} FROM objectives LIMIT 1")
        except Exception:
            try:
                conn.execute(f"ALTER TABLE objectives ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass

    # Add status column
    try:
        conn.execute("SELECT status FROM objectives LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE objectives ADD COLUMN status TEXT DEFAULT 'Not Started'")
            conn.execute("UPDATE objectives SET status = 'Complete' WHERE captured = 1")
        except Exception:
            pass

    # Migrate: add domain_id to artifacts
    try:
        conn.execute("SELECT domain_id FROM artifacts LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE artifacts ADD COLUMN domain_id INTEGER REFERENCES domains(id)")
        except Exception:
            pass

    # Migrate: add file_created to artifacts
    try:
        conn.execute("SELECT file_created FROM artifacts LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE artifacts ADD COLUMN file_created TEXT")
        except Exception:
            pass

    # Migrate: add first_name, last_name to users
    for col in ["first_name", "last_name"]:
        try:
            conn.execute(f"SELECT {col} FROM users LIMIT 1")
        except Exception:
            try:
                conn.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT DEFAULT ''")
            except Exception:
                pass

    # Migrate: add obtained_method to artifacts
    try:
        conn.execute("SELECT obtained_method FROM artifacts LIMIT 1")
    except Exception:
        try:
            conn.execute("ALTER TABLE artifacts ADD COLUMN obtained_method TEXT DEFAULT ''")
        except Exception:
            pass

    # Migrate: populate artifact_objectives junction table from artifacts.objective_id
    try:
        existing = conn.execute("SELECT COUNT(*) FROM artifact_objectives").fetchone()[0]
        if existing == 0:
            conn.execute("""
                INSERT OR IGNORE INTO artifact_objectives (artifact_id, objective_id)
                SELECT id, objective_id FROM artifacts WHERE objective_id IS NOT NULL
            """)
    except Exception:
        pass

    conn.commit()

    # Check if objectives already seeded
    count = conn.execute("SELECT COUNT(*) FROM objectives").fetchone()[0]
    if count > 0:
        # Seed discussion if missing
        if os.path.exists(XLSX_171_PATH):
            sample = conn.execute("SELECT discussion FROM objectives WHERE requirement_id = '3.1.1' LIMIT 1").fetchone()
            if sample and not sample["discussion"]:
                wb2 = load_workbook(XLSX_171_PATH, read_only=True)
                ws2 = wb2["SP 800-171"]
                for row in ws2.iter_rows(min_row=2, values_only=True):
                    req_id = str(row[2] or "").strip()
                    req_type = str(row[1] or "").strip()
                    discussion = str(row[5] or "").strip()
                    if req_id:
                        conn.execute(
                            "UPDATE objectives SET requirement_type = ?, discussion = ? WHERE requirement_id = ?",
                            (req_type, discussion, req_id)
                        )
                conn.commit()
                wb2.close()
        conn.close()
        return

    # Parse xlsx -- two passes: collect all rows, then insert
    wb = load_workbook(XLSX_PATH, read_only=True)
    ws = wb["SP800-171A"]
    current_family = ""
    current_req_id = ""
    current_sec_req = ""
    current_examine = ""
    current_interview = ""
    current_test = ""

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    reqs_with_brackets = set()
    for row in all_rows:
        ident = str(row[1] or "").strip()
        if "[" in ident:
            reqs_with_brackets.add(ident.split("[")[0])

    for row in all_rows:
        family = row[0] or current_family
        identifier = str(row[1] or "").strip()
        sort_as = str(row[2] or "").strip()
        sec_req = row[3]
        obj_text = row[4]
        examine = row[5]
        interview = row[6]
        test = row[7]

        if not identifier:
            continue

        current_family = family

        if "[" not in identifier:
            current_req_id = identifier
            current_sec_req = sec_req or ""
            current_examine = examine or ""
            current_interview = interview or ""
            current_test = test or ""
            current_sort_as = sort_as
            current_obj_text = obj_text or ""
            if identifier not in reqs_with_brackets and current_obj_text:
                obj_clean = current_obj_text
                for prefix in ("Determine if: ", "Determine If: ", "Determine if:", "Determine If:"):
                    if obj_clean.startswith(prefix):
                        obj_clean = obj_clean[len(prefix):].strip()
                        break
                conn.execute("""
                    INSERT OR IGNORE INTO objectives
                    (id, family, requirement_id, sort_as, security_requirement,
                     assessment_objective, examine, interview, test)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    identifier, current_family, identifier, sort_as,
                    current_sec_req, obj_clean,
                    current_examine, current_interview, current_test
                ))
        else:
            conn.execute("""
                INSERT OR IGNORE INTO objectives
                (id, family, requirement_id, sort_as, security_requirement,
                 assessment_objective, examine, interview, test)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                identifier, current_family, current_req_id, sort_as,
                current_sec_req, obj_text or "",
                current_examine, current_interview, current_test
            ))

    conn.commit()
    wb.close()
    print(f"Database seeded from {XLSX_PATH}")

    # Seed discussion and requirement_type from SP 800-171
    if os.path.exists(XLSX_171_PATH):
        wb2 = load_workbook(XLSX_171_PATH, read_only=True)
        ws2 = wb2["SP 800-171"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            req_id = str(row[2] or "").strip()
            req_type = str(row[1] or "").strip()
            sec_req = str(row[4] or "").strip()
            discussion = str(row[5] or "").strip()
            if req_id:
                conn.execute(
                    "UPDATE objectives SET requirement_type = ?, discussion = ? WHERE requirement_id = ?",
                    (req_type, discussion, req_id)
                )
                if sec_req:
                    sec_req_clean = re.sub(r'\[\d+\]\.?', '', sec_req).strip()
                    conn.execute(
                        "UPDATE objectives SET security_requirement = ? WHERE requirement_id = ? AND (security_requirement IS NULL OR security_requirement = '')",
                        (sec_req_clean, req_id)
                    )
        conn.commit()
        wb2.close()
        print(f"Discussion and requirement types seeded from {XLSX_171_PATH}")

    conn.close()
